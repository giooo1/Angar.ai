"""Tests for the approve action and the CSV/XLSX/JSON export endpoint.

Extractions are inserted directly into the session (rather than driven
through the upload route) so each test controls the exact canonical shape
— including a Georgian line item to prove the encoding survives.
"""

from __future__ import annotations

import io
import json
import uuid
from datetime import datetime, timezone

import pytest
from fastapi.testclient import TestClient

from angar_schema.canonical import (
    CanonicalInvoice,
    Currency,
    DocumentType,
    ExtractionMetadata,
    LineItem,
    Money,
    Party,
    PartyType,
    Script,
)
from decimal import Decimal

from backend.auth import get_current_org, get_current_user
from backend.db import get_db
from backend.main import app
from backend.models import Document, Extraction, Organization, OrganizationMember, User
from backend.auth import hash_password

GEO_DESC = "სარეკლამო კამპანია"
GEO_SELLER = "ვერტექს ლოჯისტიქს"


def _canonical_with_items() -> dict:
    inv = CanonicalInvoice(
        accepted=True,
        document_type=DocumentType.REGULAR_INVOICE,
        document_number="INV-2026-01482",
        document_currency=Currency.GEL,
        seller=Party(name=GEO_SELLER, tin="405998721", party_type=PartyType.LEGAL_ENTITY, script=Script.MKHEDRULI),
        buyer=Party(name="Imedi Clinic", tin="202249110"),
        items=[
            LineItem(
                description=GEO_DESC,
                quantity=Decimal("1"),
                unit="pc",
                unit_price=Money(amount=Decimal("1850.00"), currency=Currency.GEL),
                total=Money(amount=Decimal("1850.00"), currency=Currency.GEL),
            ),
            LineItem(
                description="site service",
                quantity=Decimal("2"),
                unit=None,
                unit_price=Money(amount=Decimal("210.00"), currency=Currency.GEL),
                total=Money(amount=Decimal("420.00"), currency=Currency.GEL),
            ),
        ],
        grand_total=Money(amount=Decimal("2270.00"), currency=Currency.GEL),
        extraction=ExtractionMetadata(
            source_filename="x.pdf",
            source_pdf_sha256="abc",
            extracted_at=datetime(2026, 1, 1, tzinfo=timezone.utc),
            model_version="m",
            prompt_version="p",
        ),
    )
    return inv.model_dump(mode="json")


def _canonical_no_items() -> dict:
    inv = CanonicalInvoice(
        accepted=True,
        document_type=DocumentType.REGULAR_INVOICE,
        document_number="EMPTY-1",
        document_currency=Currency.GEL,
        items=[],
        extraction=ExtractionMetadata(
            source_filename="x.pdf",
            source_pdf_sha256="abc",
            extracted_at=datetime(2026, 1, 1, tzinfo=timezone.utc),
            model_version="m",
            prompt_version="p",
        ),
    )
    return inv.model_dump(mode="json")


def _make_extraction(db, org: Organization, user: User, canonical: dict | None) -> Extraction:
    doc = Document(
        organization_id=org.id,
        uploaded_by_user_id=user.id,
        original_filename="invoice.pdf",
        file_sha256=uuid.uuid4().hex,
        file_size_bytes=1024,
        file_mime_type="application/pdf",
        storage_path=f"store/{uuid.uuid4().hex}.pdf",
    )
    db.add(doc)
    db.flush()
    extraction = Extraction(
        document_id=doc.id,
        status="completed" if canonical else "failed",
        prompt_version="p",
        model_version="m",
        canonical_data=canonical,
    )
    db.add(extraction)
    db.commit()
    db.refresh(extraction)
    return extraction


@pytest.fixture
def client(db_session, test_user, test_org) -> TestClient:
    def _db_override():
        yield db_session

    app.dependency_overrides[get_db] = _db_override
    app.dependency_overrides[get_current_user] = lambda: test_user
    app.dependency_overrides[get_current_org] = lambda: test_org
    yield TestClient(app)
    app.dependency_overrides.clear()


@pytest.fixture
def other_org(db_session, test_user) -> Organization:
    """A second org the test_user does NOT act as — used for cross-org 404s."""
    org = Organization(
        name="Other Org",
        monthly_extraction_quota=50,
        monthly_extractions_used=0,
        quota_reset_at=datetime.now(tz=timezone.utc),
    )
    db_session.add(org)
    db_session.flush()
    other_user = User(email="other@example.com", password_hash=hash_password("x"), full_name="O")
    db_session.add(other_user)
    db_session.flush()
    db_session.add(OrganizationMember(organization_id=org.id, user_id=other_user.id, role="owner"))
    db_session.commit()
    return org


# ---------------------------------------------------------------------------
# Approve
# ---------------------------------------------------------------------------

class TestApprove:
    def test_approve_sets_timestamp_and_returns_it(self, client, db_session, test_org, test_user) -> None:
        ex = _make_extraction(db_session, test_org, test_user, _canonical_with_items())
        r = client.post(f"/api/v1/extractions/{ex.id}/approve")
        assert r.status_code == 200, r.text
        assert r.json()["approved_at"] is not None

    def test_approve_is_idempotent(self, client, db_session, test_org, test_user) -> None:
        ex = _make_extraction(db_session, test_org, test_user, _canonical_with_items())
        first = client.post(f"/api/v1/extractions/{ex.id}/approve").json()["approved_at"]
        second = client.post(f"/api/v1/extractions/{ex.id}/approve").json()["approved_at"]
        assert first is not None and second is not None

    def test_get_extraction_reflects_approval(self, client, db_session, test_org, test_user) -> None:
        ex = _make_extraction(db_session, test_org, test_user, _canonical_with_items())
        assert client.get(f"/api/v1/extractions/{ex.id}").json()["approved_at"] is None
        client.post(f"/api/v1/extractions/{ex.id}/approve")
        assert client.get(f"/api/v1/extractions/{ex.id}").json()["approved_at"] is not None

    def test_approve_other_org_is_404(self, client, db_session, other_org, test_user) -> None:
        ex = _make_extraction(db_session, other_org, test_user, _canonical_with_items())
        r = client.post(f"/api/v1/extractions/{ex.id}/approve")
        assert r.status_code == 404


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

class TestExportCsv:
    def test_csv_headers_and_bom(self, client, db_session, test_org, test_user) -> None:
        ex = _make_extraction(db_session, test_org, test_user, _canonical_with_items())
        r = client.get(f"/api/v1/extractions/{ex.id}/export?format=csv")
        assert r.status_code == 200, r.text
        assert r.headers["content-type"].startswith("text/csv")
        assert "attachment" in r.headers["content-disposition"]
        assert r.content.startswith(b"\xef\xbb\xbf")  # UTF-8 BOM

    def test_csv_georgian_roundtrips(self, client, db_session, test_org, test_user) -> None:
        ex = _make_extraction(db_session, test_org, test_user, _canonical_with_items())
        body = client.get(f"/api/v1/extractions/{ex.id}/export?format=csv").content.decode("utf-8")
        assert GEO_DESC in body
        assert GEO_SELLER in body

    def test_csv_row_count_matches_items(self, client, db_session, test_org, test_user) -> None:
        ex = _make_extraction(db_session, test_org, test_user, _canonical_with_items())
        body = client.get(f"/api/v1/extractions/{ex.id}/export?format=csv").content.decode("utf-8-sig")
        data_lines = [ln for ln in body.splitlines() if ln.strip()]
        assert len(data_lines) == 1 + 2  # header + 2 items

    def test_csv_no_items_yields_one_row(self, client, db_session, test_org, test_user) -> None:
        ex = _make_extraction(db_session, test_org, test_user, _canonical_no_items())
        body = client.get(f"/api/v1/extractions/{ex.id}/export?format=csv").content.decode("utf-8-sig")
        data_lines = [ln for ln in body.splitlines() if ln.strip()]
        assert len(data_lines) == 1 + 1  # header + one summary row

    def test_csv_default_format(self, client, db_session, test_org, test_user) -> None:
        ex = _make_extraction(db_session, test_org, test_user, _canonical_with_items())
        r = client.get(f"/api/v1/extractions/{ex.id}/export")
        assert r.status_code == 200
        assert r.headers["content-type"].startswith("text/csv")


class TestExportXlsx:
    def test_xlsx_reopens_and_georgian_survives(self, client, db_session, test_org, test_user) -> None:
        from openpyxl import load_workbook

        ex = _make_extraction(db_session, test_org, test_user, _canonical_with_items())
        r = client.get(f"/api/v1/extractions/{ex.id}/export?format=xlsx")
        assert r.status_code == 200, r.text
        assert "spreadsheetml" in r.headers["content-type"]
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        values = [c.value for row in ws.iter_rows() for c in row]
        assert GEO_DESC in values
        assert GEO_SELLER in values


class TestExportJson:
    def test_json_deep_equals_canonical(self, client, db_session, test_org, test_user) -> None:
        canonical = _canonical_with_items()
        ex = _make_extraction(db_session, test_org, test_user, canonical)
        r = client.get(f"/api/v1/extractions/{ex.id}/export?format=json")
        assert r.status_code == 200, r.text
        assert r.headers["content-type"].startswith("application/json")
        assert json.loads(r.content) == canonical

    def test_json_keeps_georgian_literal(self, client, db_session, test_org, test_user) -> None:
        ex = _make_extraction(db_session, test_org, test_user, _canonical_with_items())
        body = client.get(f"/api/v1/extractions/{ex.id}/export?format=json").content.decode("utf-8")
        assert GEO_DESC in body  # not \uXXXX-escaped


class TestExportErrors:
    def test_unknown_format_is_422(self, client, db_session, test_org, test_user) -> None:
        # FastAPI rejects a Literal query value it doesn't recognize with 422.
        ex = _make_extraction(db_session, test_org, test_user, _canonical_with_items())
        r = client.get(f"/api/v1/extractions/{ex.id}/export?format=pdf")
        assert r.status_code == 422

    def test_export_without_canonical_is_409(self, client, db_session, test_org, test_user) -> None:
        ex = _make_extraction(db_session, test_org, test_user, None)
        r = client.get(f"/api/v1/extractions/{ex.id}/export?format=csv")
        assert r.status_code == 409

    def test_export_other_org_is_404(self, client, db_session, other_org, test_user) -> None:
        ex = _make_extraction(db_session, other_org, test_user, _canonical_with_items())
        r = client.get(f"/api/v1/extractions/{ex.id}/export?format=csv")
        assert r.status_code == 404
